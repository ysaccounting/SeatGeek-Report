"""
SG2 Report — backend.

Upload one or more **Invoice Details** exports and one or more **Purchase Details**
exports (.csv / .xlsx / .xlsm). Produces the monthly SG2 workbook with three tabs:

  * Summary          — YTD P&L by Inventory Type (NFL/MLB/NBA/NHL/Concerts/Other),
                       built from live formulas that reference the Invoice Details tab.
  * Invoice Details  — every invoice row, with a `Category` column prepended (col A).
  * Purchase Details — every purchase row, passed through as-is.

The only thing the app decides is the `Category` for each invoice row. Everything on
the Summary is a live Excel formula (SUMIF / SUMIFS) so the workbook stays auditable and
recalculates if a category is edited by hand.

Categorisation runs on the Performer/Team value. The four major-league rosters are matched
exactly; everything else is decided by signals already in the row (the Performer/Opponent
column and the TextTags). A persistent overrides file records any human decisions so the
same name is never reviewed twice. Anything the rules genuinely cannot place is surfaced
for a quick human check *before* the final workbook is produced (the review gate).
"""

import io
import os
import re
import csv
import json
import time
import uuid
import pickle
import shutil
import zipfile
import tempfile
import datetime as dt
import collections
from urllib.parse import quote

import pandas as pd
from flask import Flask, request, jsonify, send_file, send_from_directory, abort
from openpyxl import Workbook
from openpyxl.cell import WriteOnlyCell
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

app = Flask(__name__, static_folder=None)
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
STORE_DIR = os.path.join(tempfile.gettempdir(), "sg2_store")
os.makedirs(STORE_DIR, exist_ok=True)

# Overrides persist *human* category decisions (a name you explicitly changed in the
# review gate). They win over the rules. The "seen" set records every performer the app
# has already processed, so the review gate only surfaces genuinely new names. Point both
# at a mounted Railway volume for durability across redeploys (see README).
OVERRIDES_PATH = os.environ.get("OVERRIDES_PATH", os.path.join(BASE_DIR, "overrides.json"))
SEEN_PATH = os.environ.get("SEEN_PATH", os.path.join(BASE_DIR, "seen.json"))


# =========================================================================== #
# DEAL TERMS & CONFIG — edit these to tune the report
# =========================================================================== #

DEAL_CODE = "SG2"
REPORT_NAME = "SeatGeek Report"   # output file is "{REPORT_NAME} - {Month} {Year}.xlsx"
PROFIT_SHARE = 0.30          # SG share of profit
FEE_RATE = 0.07              # fee as % of revenue, taken before the profit share
CAPITAL_INVESTED = 19881500  # "Capital invested by SG to date" (fixed; edit when it changes)

# Marketplace columns on the Summary. (label, exact Client value). Everything not
# listed here rolls into "% Sales Other". Excel SUMIFS matching is case-insensitive.
MARKETPLACES = [("SG", "SeatGeek"), ("VS", "Vivid Seats"), ("SH", "StubHub")]

CATEGORY_ORDER = ["NFL", "MLB", "NBA", "NHL", "Concerts", "Other"]

# Column names the app relies on (matched case-insensitively, trimmed).
INV_PERFORMER = "Performer/Team"
INV_OPPONENT = "Performer/Opponent"
INV_TAGS = "TextTags"
INV_CANCELLED = "Cancelled"
INV_CLIENT = "Client"
INV_PRICE = "Total Price"
INV_COST = "Total Cost"
INV_EVENT_DATE = "Event Date"
INV_VENUE = "Venue"
PUR_TOTAL_COST = "Total Cost"


# --------------------------------------------------------------------------- #
# League rosters (exact-match). Include recent renames / relocations.
# --------------------------------------------------------------------------- #
NFL = {"Arizona Cardinals","Atlanta Falcons","Baltimore Ravens","Buffalo Bills","Carolina Panthers","Chicago Bears","Cincinnati Bengals","Cleveland Browns","Dallas Cowboys","Denver Broncos","Detroit Lions","Green Bay Packers","Houston Texans","Indianapolis Colts","Jacksonville Jaguars","Kansas City Chiefs","Las Vegas Raiders","Los Angeles Chargers","Los Angeles Rams","Miami Dolphins","Minnesota Vikings","New England Patriots","New Orleans Saints","New York Giants","New York Jets","Philadelphia Eagles","Pittsburgh Steelers","San Francisco 49ers","Seattle Seahawks","Tampa Bay Buccaneers","Tennessee Titans","Washington Commanders","Washington Football Team","Oakland Raiders","San Diego Chargers"}
MLB = {"Arizona Diamondbacks","Atlanta Braves","Baltimore Orioles","Boston Red Sox","Chicago Cubs","Chicago White Sox","Cincinnati Reds","Cleveland Guardians","Cleveland Indians","Colorado Rockies","Detroit Tigers","Houston Astros","Kansas City Royals","Los Angeles Angels","Los Angeles Dodgers","Miami Marlins","Milwaukee Brewers","Minnesota Twins","New York Mets","New York Yankees","Athletics","Oakland Athletics","Philadelphia Phillies","Pittsburgh Pirates","San Diego Padres","San Francisco Giants","Seattle Mariners","St. Louis Cardinals","Tampa Bay Rays","Texas Rangers","Toronto Blue Jays","Washington Nationals"}
NBA = {"Atlanta Hawks","Boston Celtics","Brooklyn Nets","Charlotte Hornets","Chicago Bulls","Cleveland Cavaliers","Dallas Mavericks","Denver Nuggets","Detroit Pistons","Golden State Warriors","Houston Rockets","Indiana Pacers","LA Clippers","Los Angeles Clippers","Los Angeles Lakers","Memphis Grizzlies","Miami Heat","Milwaukee Bucks","Minnesota Timberwolves","New Orleans Pelicans","New York Knicks","Oklahoma City Thunder","Orlando Magic","Philadelphia 76ers","Phoenix Suns","Portland Trail Blazers","Sacramento Kings","San Antonio Spurs","Toronto Raptors","Utah Jazz","Washington Wizards"}
NHL = {"Anaheim Ducks","Boston Bruins","Buffalo Sabres","Calgary Flames","Carolina Hurricanes","Chicago Blackhawks","Colorado Avalanche","Columbus Blue Jackets","Dallas Stars","Detroit Red Wings","Edmonton Oilers","Florida Panthers","Los Angeles Kings","Minnesota Wild","Montreal Canadiens","Montréal Canadiens","Nashville Predators","New Jersey Devils","New York Islanders","New York Rangers","Ottawa Senators","Philadelphia Flyers","Pittsburgh Penguins","San Jose Sharks","Seattle Kraken","St. Louis Blues","Tampa Bay Lightning","Toronto Maple Leafs","Utah Mammoth","Utah Hockey Club","Vancouver Canucks","Vegas Golden Knights","Washington Capitals","Winnipeg Jets","Arizona Coyotes"}
LEAGUE_SETS = [("NFL", NFL), ("MLB", MLB), ("NBA", NBA), ("NHL", NHL)]
LEAGUE_ALL = NFL | MLB | NBA | NHL

_THEATER_RE = re.compile(r"the musical|- musical|a new musical|\(play\)|\(musical\)| on stage|the play\b", re.I)

# Known stage productions -> Other (matched via theater_core, prefix-aware so subtitles
# like "A Beautiful Noise - Music of Neil Diamond" still match "a beautiful noise").
# Editable: add a show here and it leaves the Concerts bucket immediately.
THEATER_TITLES = {
    "hamilton", "wicked", "the phantom of the opera", "phantom of the opera", "les miserables",
    "the lion king", "hadestown", "the book of mormon", "moulin rouge", "aladdin",
    "mj", "six", "back to the future", "the great gatsby", "a beautiful noise",
    "water for elephants", "harry potter and the cursed child", "the wiz", "suffs",
    "hells kitchen", "beauty and the beast", "the outsiders", "cabaret", "sweeney todd",
    "merrily we roll along", "the notebook", "kimberly akimbo", "parade", "funny girl",
    "some like it hot", "mrs doubtfire", "dear evan hansen", "come from away",
    "to kill a mockingbird", "mamma mia", "the wizard of oz", "mean girls", "juliet",
    "tina", "aint too proud", "jagged little pill", "stranger things", "death becomes her",
    "the great comet", "shucked", "spamalot", "hairspray", "jersey boys", "a streetcar named desire",
    "the play that goes wrong", "matilda", "pretty woman", "the addams family", "waitress",
    "riverdance", "stomp", "blue man group", "the nutcracker", "swan lake", "moulin rouge",
}

# Obvious non-concert live events / non-team sports -> Other (incl. wrestling & MMA).
SHOW_EVENT_RE = re.compile(
    r"\bon ice\b|cirque|globetrotters|monster jam|disney on|paw patrol|sesame street|"
    r"\brodeo\b|nascar|grand prix|\bwwe\b|\bufc\b|\baew\b|bellator|\bpfl\b|\bnxt\b|"
    r"wrestlemania|summerslam|royal rumble|smackdown|marathon|"
    r"savannah bananas|party animals|monster energy|supercross|\bprca\b|stars on ice|"
    r"dancing with the stars|\bwnba\b|harlem", re.I)


def theater_core(name):
    """Normalized stage-title key: strip 'tickets', then trailing show words."""
    s = normalize_name(name)
    s = re.sub(r"\s+(the musical|musical|the play|play|on ice|on stage|live|tour)$", "", s).strip()
    s = re.sub(r"[^a-z0-9 ]", "", s)
    return re.sub(r"\s+", " ", s).strip()


def is_theater_title(name):
    """True if the performer is a known stage production (title or 'X - subtitle')."""
    core = theater_core(name)
    if not core:
        return False
    return any(core == t or core.startswith(t + " ") for t in THEATER_TITLES)


def normalize_name(s):
    """Lookup key: lowercase, trim, drop trailing 'tickets'/'parking', collapse spaces.
    Conservative on purpose — merges 'Noah Kahan Tickets' with 'Noah Kahan' without
    risking collisions between distinct teams."""
    s = str(s).strip().lower()
    s = re.sub(r"\s+(tickets?|parking)$", "", s)
    return re.sub(r"\s+", " ", s).strip()


def canon_category(x):
    s = str(x).strip().lower()
    return {"nfl": "NFL", "mlb": "MLB", "nba": "NBA", "nhl": "NHL",
            "concert": "Concerts", "concerts": "Concerts", "other": "Other"}.get(s, "Other")


# =========================================================================== #
# File reading
# =========================================================================== #

def _read_any(filename, data):
    """Read a CSV/XLSX/XLSM upload into a DataFrame, inferring dtypes so amount
    columns stay numeric. For workbooks, the sheet with the most cells wins."""
    low = filename.lower()
    if low.endswith(".csv"):
        text = data.decode("utf-8-sig", errors="replace")
        return pd.read_csv(io.StringIO(text), on_bad_lines="skip")
    sheets = pd.read_excel(io.BytesIO(data), sheet_name=None)
    if not sheets:
        return pd.DataFrame()
    return max(sheets.values(), key=lambda d: max(d.shape[0], 1) * max(d.shape[1], 1))


def _concat(frames):
    """Concatenate frames, preserving first-seen column order."""
    frames = [f for f in frames if f is not None and not f.empty]
    if not frames:
        return pd.DataFrame()
    order, seen = [], set()
    for f in frames:
        for c in f.columns:
            if c not in seen:
                seen.add(c); order.append(c)
    out = pd.concat(frames, ignore_index=True)
    return out.reindex(columns=order)


def _resolve(df, wanted):
    """Find a column by case-insensitive / trimmed name. Returns actual name or None."""
    norm = {str(c).strip().lower(): c for c in df.columns}
    return norm.get(wanted.strip().lower())


# =========================================================================== #
# Categorisation
# =========================================================================== #

def _tag_counts(series):
    c = collections.Counter()
    for t in series.dropna():
        for p in str(t).split(","):
            p = p.strip().upper()
            if p:
                c[p] += 1
    return c


def classify(name, sub, opp_col, tag_col):
    """Return (category, reason) for one performer and its rows. Rule-based, using
    knowledge first (rosters, stage shows, live events), then the opponent signal,
    then TextTags if present, then a Concerts default for a lone act."""
    if name in LEAGUE_ALL:
        for lg, s in LEAGUE_SETS:
            if name in s:
                return lg, "league_exact"
    nl = name.lower()
    for lg, s in LEAGUE_SETS:                       # league name w/ suffix (parking, placeholder)
        if any(t.lower() in nl for t in s):
            return lg, "league_substr"
    if is_theater_title(name):
        return "Other", "theater_title"
    if SHOW_EVENT_RE.search(name) or _THEATER_RE.search(name):
        return "Other", "show_event"
    has_opp = sub[opp_col].notna().any() if opp_col is not None else False
    if has_opp:
        return "Other", "has_opponent"          # a matchup that isn't a major league = other sports
    tc = _tag_counts(sub[tag_col]) if tag_col is not None else collections.Counter()
    n = len(sub)
    b, c, s = tc.get("BROADWAY", 0), tc.get("CONCERT", 0), tc.get("SPORTS", 0)
    if b >= 0.5 * n and b > 0:
        return "Other", "broadway_tag"
    if s > c and s > 0:
        return "Other", "sports_tag"
    return "Concerts", "default_concert"         # lone act -> concert (the gray call)


def build_mapping(inv_df, overrides):
    """Map each unique Performer/Team -> (category, reason). Overrides win."""
    perf = _resolve(inv_df, INV_PERFORMER)
    opp = _resolve(inv_df, INV_OPPONENT)
    tag = _resolve(inv_df, INV_TAGS)
    if perf is None:
        raise ValueError(f"Invoice files are missing a '{INV_PERFORMER}' column.")
    s = inv_df[perf].apply(lambda x: str(x).strip() if pd.notna(x) else "")
    work = inv_df.assign(_pt=s)
    mapping, reasons, counts = {}, {}, {}
    for name, grp in work.groupby("_pt"):
        counts[name] = len(grp)
        if name == "":
            mapping[name], reasons[name] = "Other", "blank"
            continue
        key = normalize_name(name)
        if key in overrides:
            mapping[name], reasons[name] = overrides[key], "lookup"
            continue
        cat, why = classify(name, grp, opp, tag)
        mapping[name], reasons[name] = cat, why
    return mapping, reasons, counts


# =========================================================================== #
# Workbook builder (write-only for memory safety on large detail sheets)
# =========================================================================== #

_MONEY2 = '"$"#,##0.00'
_MONEY0 = '"$"#,##0'
_PCT = '0.00%'
_PCT0 = '0%'
_HDR_FILL = PatternFill("solid", fgColor="374151")
_TOT_FILL = PatternFill("solid", fgColor="EAECF0")


def _woc(ws, value, *, bold=False, white=False, fill=None, numfmt=None, wrap=False):
    c = WriteOnlyCell(ws, value=value)
    if bold or white:
        c.font = Font(bold=bold, color="FFFFFF" if white else "111111")
    if fill:
        c.fill = fill
    if numfmt:
        c.number_format = numfmt
    if wrap:
        c.alignment = Alignment(wrap_text=True, vertical="center")
    return c


def build_report(inv_df, purchase_df, mapping, report_label, category_list=None,
                 summary_title="Year to date data from deal start"):
    """Return xlsx bytes for the four-tab workbook (Summary, Invoice Details, Purchase
    Details, Category). `category_list` is the running master list as (performer,
    category) pairs; `summary_title` is the A1 date-range line."""
    perf = _resolve(inv_df, INV_PERFORMER)
    inv_cols = list(inv_df.columns)
    out_cols = ["Category"] + inv_cols

    # Column letters the Summary formulas need (Category is forced to col A).
    def letter(colname, cols, offset):
        actual = _resolve_in(cols, colname)
        if actual is None:
            return None
        return get_column_letter(cols.index(actual) + 1 + offset)

    price_L = letter(INV_PRICE, inv_cols, 1)
    cost_L = letter(INV_COST, inv_cols, 1)
    client_L = letter(INV_CLIENT, inv_cols, 1)
    if not (price_L and cost_L and client_L):
        missing = [n for n, L in [(INV_PRICE, price_L), (INV_COST, cost_L), (INV_CLIENT, client_L)] if not L]
        raise ValueError("Invoice files are missing required column(s): " + ", ".join(missing))

    pur_cols = list(purchase_df.columns)
    pur_cost_actual = _resolve_in(pur_cols, PUR_TOTAL_COST)
    pur_cost_L = get_column_letter(pur_cols.index(pur_cost_actual) + 1) if pur_cost_actual else None

    wb = Workbook(write_only=True)

    # ---------------- Summary ----------------
    sm = wb.create_sheet("Summary")
    # Widths sized so every header sits on one line (no wrapping).
    col_widths = {"A": 16, "B": 19, "C": 14, "D": 14, "E": 14,
                  "F": 14, "G": 15, "H": 13, "I": 21, "J": 31}
    for col, w in col_widths.items():
        sm.column_dimensions[col].width = w
    sm.append([_woc(sm, summary_title, bold=True)])
    headers = ["Inventory Type", "Sales within deal",
               "% Sales on SG", "% Sales on VS", "% Sales on SH", "% Sales Other",
               "COGS", "Profit", "Profit Share to SG", "Profit Share to SG Less fees"]
    sm.append([_woc(sm, h, bold=True, white=True, fill=_HDR_FILL) for h in headers])

    first, last = 3, 3 + len(CATEGORY_ORDER) - 1
    for i, cat in enumerate(CATEGORY_ORDER):
        r = first + i
        row = [_woc(sm, cat, bold=True)]
        row.append(_woc(sm, f"=SUMIF('Invoice Details'!$A:$A,Summary!$A{r},'Invoice Details'!${price_L}:${price_L})", numfmt=_MONEY0))
        for _, client in MARKETPLACES:
            row.append(_woc(sm, f'=SUMIFS(\'Invoice Details\'!${price_L}:${price_L},\'Invoice Details\'!${client_L}:${client_L},"{client}",\'Invoice Details\'!$A:$A,Summary!$A{r})/$B{r}', numfmt=_PCT0))
        c0 = get_column_letter(3)
        cN = get_column_letter(2 + len(MARKETPLACES))
        row.append(_woc(sm, f"=1-SUM({c0}{r}:{cN}{r})", numfmt=_PCT0))
        row.append(_woc(sm, f"=SUMIF('Invoice Details'!$A:$A,Summary!$A{r},'Invoice Details'!${cost_L}:${cost_L})", numfmt=_MONEY0))
        row.append(_woc(sm, f"=B{r}-G{r}", numfmt=_MONEY0))
        row.append(_woc(sm, f"=H{r}*{PROFIT_SHARE}", numfmt=_MONEY0))
        row.append(_woc(sm, f"=(H{r}-{FEE_RATE}*B{r})*{PROFIT_SHARE}", numfmt=_MONEY0))
        sm.append(row)

    # Totals
    tr = [_woc(sm, "Totals", bold=True, fill=_TOT_FILL)]
    def tcell(formula, fmt):
        return _woc(sm, formula, bold=True, fill=_TOT_FILL, numfmt=fmt)
    tr.append(tcell(f"=SUM(B{first}:B{last})", _MONEY0))
    for col in "CDEF":
        tr.append(tcell(f"=SUMPRODUCT($B${first}:$B${last}*{col}{first}:{col}{last})/$B${last+1}", _PCT0))
    for col in "GHIJ":
        tr.append(tcell(f"=SUM({col}{first}:{col}{last})", _MONEY0))
    sm.append(tr)

    sm.append([])
    sm.append([])
    sm.append([_woc(sm, "Capital invested by SG to date", bold=True),
               _woc(sm, CAPITAL_INVESTED, numfmt=_MONEY0)])
    fund = f"=SUM('Purchase Details'!{pur_cost_L}:{pur_cost_L})" if pur_cost_L else 0
    sm.append([_woc(sm, "Size of inventory Fund", bold=True),
               _woc(sm, fund, numfmt=_MONEY0)])

    # ---------------- Invoice Details ----------------
    inv = wb.create_sheet("Invoice Details")
    inv.append([_woc(inv, h, bold=True, white=True, fill=_HDR_FILL) for h in out_cols])
    perf_idx = inv_cols.index(perf)
    for rec in inv_df.itertuples(index=False, name=None):
        name = str(rec[perf_idx]).strip() if rec[perf_idx] is not None and not _isnan(rec[perf_idx]) else ""
        cat = mapping.get(name, "Other")
        row = [cat] + [None if _isnan(v) else v for v in rec]
        inv.append(row)

    # ---------------- Purchase Details ----------------
    pur = wb.create_sheet("Purchase Details")
    pur.append([_woc(pur, h, bold=True, white=True, fill=_HDR_FILL) for h in pur_cols])
    for rec in purchase_df.itertuples(index=False, name=None):
        pur.append([None if _isnan(v) else v for v in rec])

    # ---------------- Category (running master list, LAST tab) ----------------
    # Same shape as the uploaded master (Performer/Team, League) so this tab can be
    # fed back in next month. Sorted by category A-Z, then performer A-Z.
    cat_sheet = wb.create_sheet("Category")
    cat_sheet.column_dimensions["A"].width = 48
    cat_sheet.column_dimensions["B"].width = 14
    cat_sheet.append([_woc(cat_sheet, h, bold=True, white=True, fill=_HDR_FILL)
                      for h in ("Performer/Team", "League")])
    if category_list is None:
        category_list = [(name, cat) for name, cat in mapping.items() if str(name).strip()]
    cat_rows = sorted(((str(n).strip(), c) for n, c in category_list if str(n).strip()),
                      key=lambda t: (str(t[1]).lower(), str(t[0]).lower()))
    for name, cat in cat_rows:
        cat_sheet.append([_woc(cat_sheet, name), _woc(cat_sheet, cat)])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _isnan(v):
    try:
        return v is None or (isinstance(v, float) and pd.isna(v))
    except Exception:
        return False


def _resolve_in(cols, wanted):
    norm = {str(c).strip().lower(): c for c in cols}
    return norm.get(wanted.strip().lower())


def read_category_list(files):
    """Read an uploaded master category list into {normkey: (display_name, category)}.
    Accepts the standalone list (Performer/Team + League) or any report/sheet that has
    a performer column and a League / Category / Inventory Type column."""
    master = {}
    for name, data in files:
        try:
            if name.lower().endswith(".csv"):
                sheets = {"_": pd.read_csv(io.StringIO(data.decode("utf-8-sig", errors="replace")))}
            else:
                sheets = pd.read_excel(io.BytesIO(data), sheet_name=None)
        except Exception:
            continue
        for df in sheets.values():
            pcol = _resolve(df, INV_PERFORMER) or _resolve(df, "Performer")
            ccol = next((c for c in (_resolve(df, "League"), _resolve(df, "Category"),
                                     _resolve(df, "Inventory Type")) if c), None)
            if pcol is None or ccol is None:
                continue
            for p, c in zip(df[pcol], df[ccol]):
                if pd.notna(p) and str(p).strip() and pd.notna(c):
                    k = normalize_name(p)
                    if k:
                        master[k] = (str(p).strip(), canon_category(c))
    return master


def merge_master(prior_master, mapping):
    """Carry the prior master forward and fold in this month's performers/categories.
    Returns a list of (display_name, category)."""
    master = dict(prior_master)
    for name, cat in mapping.items():
        if str(name).strip():
            master[normalize_name(name)] = (str(name).strip(), cat)
    return [val for val in master.values()]


def build_category_file(category_list):
    """Standalone running-list workbook (one Category sheet) for easy re-upload."""
    wb = Workbook(write_only=True)
    ws = wb.create_sheet("Category")
    ws.column_dimensions["A"].width = 48
    ws.column_dimensions["B"].width = 14
    ws.append([_woc(ws, h, bold=True, white=True, fill=_HDR_FILL) for h in ("Performer/Team", "League")])
    for name, cat in sorted(((str(n).strip(), c) for n, c in category_list if str(n).strip()),
                            key=lambda t: (str(t[1]).lower(), str(t[0]).lower())):
        ws.append([_woc(ws, name), _woc(ws, cat)])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# =========================================================================== #
# Overrides + housekeeping
# =========================================================================== #

def load_overrides():
    try:
        with open(OVERRIDES_PATH, "r", encoding="utf-8") as fh:
            d = json.load(fh)
            return {str(k): str(v) for k, v in d.items()}
    except (OSError, ValueError):
        return {}


def save_overrides(new_pairs):
    cur = load_overrides()
    cur.update({normalize_name(k): canon_category(v) for k, v in new_pairs.items()})
    try:
        with open(OVERRIDES_PATH, "w", encoding="utf-8") as fh:
            json.dump(cur, fh, indent=2, sort_keys=True)
    except OSError:
        pass
    return cur


def seed_overrides_from_df(df, performer_col=None, category_col="Category"):
    """Populate the lookup from a previously-categorised report (a sheet with a
    Performer/Team column and a Category column). Existing keys are preserved
    unless overwritten. Returns count of keys added/updated."""
    pcol = performer_col or _resolve(df, INV_PERFORMER)
    ccol = _resolve(df, category_col)
    if pcol is None or ccol is None:
        raise ValueError("Seed file needs both a Performer/Team and a Category column.")
    sub = df[[pcol, ccol]].dropna()
    pairs = {}
    for name, cat in zip(sub[pcol], sub[ccol]):
        key = normalize_name(name)
        if key:
            pairs[key] = canon_category(cat)
    save_overrides(pairs)
    return len(pairs)


def load_seen():
    try:
        with open(SEEN_PATH, "r", encoding="utf-8") as fh:
            return set(json.load(fh))
    except (OSError, ValueError):
        return set()


def add_seen(keys):
    seen = load_seen()
    seen.update(normalize_name(k) for k in keys if str(k).strip())
    try:
        with open(SEEN_PATH, "w", encoding="utf-8") as fh:
            json.dump(sorted(seen), fh, indent=0)
    except OSError:
        pass
    return seen


def _cleanup_old(max_age=12 * 3600):
    now = time.time()
    for name in os.listdir(STORE_DIR):
        p = os.path.join(STORE_DIR, name)
        try:
            if os.path.isdir(p) and now - os.path.getmtime(p) > max_age:
                shutil.rmtree(p, ignore_errors=True)
        except OSError:
            pass


def _safe(s):
    return re.sub(r'[\\/:*?"<>|]+', " ", str(s)).strip() if s else s


def _period_end(inv_df, pur_df=None):
    # Report period ends at the latest transaction in the files: invoice "Created Date"
    # and purchase "PO Created". (Event Date is ignored — it runs into the future.)
    cands = []
    icol = _resolve(inv_df, "Created Date")
    if icol is not None:
        d = pd.to_datetime(inv_df[icol], errors="coerce")
        if d.notna().any():
            cands.append(d.max())
    if pur_df is not None and not pur_df.empty:
        pcol = _resolve(pur_df, "PO Created")
        if pcol is not None:
            d = pd.to_datetime(pur_df[pcol], errors="coerce")
            if d.notna().any():
                cands.append(d.max())
    today = pd.Timestamp.now()
    when = max(cands) if cands else None
    if when is None or when > today:
        when = today.normalize().replace(day=1) - pd.Timedelta(days=1)  # prior month-end
    return when


def _invoice_period_end(inv_df):
    # A1's reporting period closes at the latest INVOICE (sales) date. A purchase PO
    # dated later can still push the *filename* month forward, but the stated sales
    # range ends with the last sale.
    today = pd.Timestamp.now()
    when = None
    icol = _resolve(inv_df, "Created Date")
    if icol is not None:
        d = pd.to_datetime(inv_df[icol], errors="coerce")
        if d.notna().any():
            when = d.max()
    if when is None or when > today:
        when = today.normalize().replace(day=1) - pd.Timedelta(days=1)  # prior month-end
    return when


def _ordinal(n):
    return f"{n}{'th' if 10 <= n % 100 <= 20 else {1: 'st', 2: 'nd', 3: 'rd'}.get(n % 10, 'th')}"


def _default_label(inv_df, pur_df=None):
    return f"{REPORT_NAME} - {_period_end(inv_df, pur_df).strftime('%B %Y')}"


def _summary_title(inv_df, pur_df=None):
    """A1 text: deal start (Jan 1st, 2026) through the end of the latest sales month."""
    when = _invoice_period_end(inv_df)
    month_end = when.replace(day=1) + pd.offsets.MonthEnd(1)
    return (f"Jan 1st, 2026 through {month_end.strftime('%B')} "
            f"{_ordinal(int(month_end.day))}, {int(month_end.year)}")


# =========================================================================== #
# Pipeline glue
# =========================================================================== #

def _prepare(invoice_files, purchase_files):
    """Read + concat + drop cancelled invoices. Returns (inv_df, purchase_df)."""
    inv = _concat([_read_any(n, d) for n, d in invoice_files])
    pur = _concat([_read_any(n, d) for n, d in purchase_files])
    if inv.empty:
        raise ValueError("No invoice rows found in the uploaded files.")
    canc = _resolve(inv, INV_CANCELLED)
    if canc is not None:
        keep = ~inv[canc].apply(lambda v: str(v).strip().lower() in ("yes", "true", "y", "1"))
        inv = inv[keep].reset_index(drop=True)
    return inv, pur


def _store(token, inv, pur, mapping, label, prior_master):
    folder = os.path.join(STORE_DIR, token)
    os.makedirs(folder, exist_ok=True)
    with open(os.path.join(folder, "data.pkl"), "wb") as fh:
        pickle.dump({"inv": inv, "pur": pur, "mapping": mapping, "label": label,
                     "prior_master": prior_master}, fh)
    return folder


def _load(token):
    with open(os.path.join(STORE_DIR, os.path.basename(token), "data.pkl"), "rb") as fh:
        return pickle.load(fh)


def _write_final(token, inv, pur, mapping, label, prior_master):
    category_list = merge_master(prior_master, mapping)
    report = build_report(inv, pur, mapping, label, category_list, _summary_title(inv, pur))
    cat_bytes = build_category_file(category_list)
    folder = os.path.join(STORE_DIR, token)
    os.makedirs(folder, exist_ok=True)
    report_fn = f"{_safe(label)}.xlsx"
    cat_fn = f"{_safe(label).replace('Report', 'Category List')}.xlsx"
    if cat_fn == report_fn:
        cat_fn = f"{_safe(label)} - Category List.xlsx"
    with open(os.path.join(folder, report_fn), "wb") as fh:
        fh.write(report)
    with open(os.path.join(folder, cat_fn), "wb") as fh:
        fh.write(cat_bytes)
    # One bundle for a single download button.
    zip_fn = f"{_safe(label)}.zip"
    with zipfile.ZipFile(os.path.join(folder, zip_fn), "w", zipfile.ZIP_DEFLATED) as zf:
        zf.writestr(report_fn, report)
        zf.writestr(cat_fn, cat_bytes)
    return zip_fn


def _stats(inv, pur, mapping):
    perf = _resolve(inv, INV_PERFORMER)
    s = inv[perf].apply(lambda x: str(x).strip() if pd.notna(x) else "")
    cats = s.map(lambda n: mapping.get(n, "Other"))
    by = {c: int((cats == c).sum()) for c in CATEGORY_ORDER}
    return {"invoice_rows": int(len(inv)), "purchase_rows": int(len(pur)),
            "by_category": by}


# =========================================================================== #
# Routes
# =========================================================================== #

@app.route("/")
def index():
    return send_from_directory(BASE_DIR, "index.html")


@app.route("/process", methods=["POST"])
def process():
    inv_files = [(f.filename, f.read()) for f in request.files.getlist("invoice") if f.filename]
    pur_files = [(f.filename, f.read()) for f in request.files.getlist("purchase") if f.filename]
    prior_files = [(f.filename, f.read()) for f in request.files.getlist("prior") if f.filename]
    if not inv_files:
        return jsonify({"error": "Please upload at least one Invoice Details file."}), 400

    try:
        inv, pur = _prepare(inv_files, pur_files)
        prior_master = read_category_list(prior_files)          # {normkey: (display, category)}
        overrides = load_overrides()
        # The uploaded running list is authoritative; persistent overrides fill any gaps.
        lookup = {**overrides, **{k: cat for k, (disp, cat) in prior_master.items()}}
        mapping, reasons, counts = build_mapping(inv, lookup)
    except Exception as exc:
        return jsonify({"error": f"{type(exc).__name__}: {exc}"}), 500

    label = (request.form.get("label") or "").strip() or _default_label(inv, pur)
    token = uuid.uuid4().hex
    _store(token, inv, pur, mapping, label, prior_master)
    _cleanup_old()

    warnings = []
    if not pur_files:
        warnings.append("No Purchase Details uploaded — \u201cSize of inventory Fund\u201d will be $0.")
    if not prior_files:
        warnings.append("No prior Category List uploaded — every performer will look new this run.")

    # Review only NEW names (not on the running list / overrides / seen) whose rule guess
    # is the gray Concerts default. The most common venue is shown as a theatrical clue.
    known = set(prior_master) | set(overrides) | load_seen()
    review_names = [n for n, why in reasons.items()
                    if why == "default_concert" and normalize_name(n) not in known]
    venues = {}
    perf = _resolve(inv, INV_PERFORMER)
    vcol = _resolve(inv, INV_VENUE)
    if review_names and vcol is not None:
        names_set = set(review_names)
        key = inv[perf].apply(lambda x: str(x).strip() if pd.notna(x) else "")
        sub = inv[key.isin(names_set)].assign(_pt=key[key.isin(names_set)])
        for nm, grp in sub.groupby("_pt"):
            vc = grp[vcol].dropna().astype(str)
            venues[nm] = vc.value_counts().index[0] if not vc.empty else ""
    review = sorted(
        [{"performer": n, "suggested": mapping[n], "rows": counts.get(n, 0),
          "venue": venues.get(n, "")} for n in review_names],
        key=lambda d: -d["rows"])

    if review:
        return jsonify({"status": "review", "token": token, "label": label,
                        "items": review, "warnings": warnings})

    add_seen(mapping.keys())
    zip_fn = _write_final(token, inv, pur, mapping, label, prior_master)
    return jsonify({"status": "ready", "token": token, "filename": zip_fn,
                    "download_url": f"/download/{token}?f={quote(zip_fn)}",
                    "label": label, "warnings": warnings, "stats": _stats(inv, pur, mapping)})


@app.route("/finalize", methods=["POST"])
def finalize():
    body = request.get_json(force=True, silent=True) or {}
    token = body.get("token", "")
    corrections = {str(k).strip(): str(v).strip() for k, v in (body.get("corrections") or {}).items()
                   if str(v).strip() in CATEGORY_ORDER}
    try:
        st = _load(token)
    except (OSError, pickle.PickleError):
        return jsonify({"error": "Session expired — please re-upload the files."}), 404

    mapping = st["mapping"]
    # Only names you actually CHANGED become locked overrides; accepting a rule guess
    # leaves the name rule-governed. The reviewed names also fold into the running list.
    changed = {p: v for p, v in corrections.items() if v != mapping.get(p)}
    mapping.update(corrections)
    if changed:
        save_overrides(changed)
    add_seen(mapping.keys())

    try:
        zip_fn = _write_final(token, st["inv"], st["pur"], mapping, st["label"],
                              st.get("prior_master", {}))
    except Exception as exc:
        return jsonify({"error": f"{type(exc).__name__}: {exc}"}), 500

    return jsonify({"status": "ready", "token": token, "filename": zip_fn,
                    "download_url": f"/download/{token}?f={quote(zip_fn)}",
                    "label": st["label"], "stats": _stats(st["inv"], st["pur"], mapping)})


@app.route("/download/<token>")
def download(token):
    folder = os.path.join(STORE_DIR, os.path.basename(token))
    if not os.path.isdir(folder):
        abort(404)
    want = request.args.get("f")
    files = [f for f in os.listdir(folder) if f.lower().endswith((".xlsx", ".zip"))]
    if not files:
        abort(404)
    if want and want in files:
        pick = want
    else:
        zips = [f for f in files if f.lower().endswith(".zip")]
        pick = zips[0] if zips else sorted(files, key=len)[0]
    mime = ("application/zip" if pick.lower().endswith(".zip")
            else "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    return send_file(os.path.join(folder, pick), mimetype=mime,
                     as_attachment=True, download_name=pick)


if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    app.run(host="0.0.0.0", port=port, debug=True)
